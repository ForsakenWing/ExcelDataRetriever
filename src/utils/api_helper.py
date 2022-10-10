from openpyxl import load_workbook
from .excel_parser import workbook_serializer
from fastapi import UploadFile, HTTPException
from tempfile import TemporaryFile


async def files_handler(files: list[UploadFile]) -> list:
    if len(files) <= 1:
        try:
            tmp_file = TemporaryFile()
            content = await files[0].read()
            tmp_file.write(content)
            wb = load_workbook(tmp_file, read_only=True)
        except Exception as e:
            raise HTTPException(status_code=401, detail=f"{e}")
        else:
            return [{1: workbook_serializer(wb)}]
    result = []
    for index, file in enumerate(files, start=1):
        try:
            tmp_file = TemporaryFile()
            content = await file.read()
            tmp_file.write(content)
            wb = load_workbook(tmp_file, read_only=True)
        except Exception as e:
            result.append({index: f"{e}"})
        else:
            result.append({index: workbook_serializer(wb)})
    return result
    
    