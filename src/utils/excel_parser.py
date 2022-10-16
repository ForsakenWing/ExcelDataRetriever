from openpyxl import load_workbook, Workbook
from openpyxl.utils.exceptions import InvalidFileException
from .cli import parser
import os
import logging
from typing import Literal
from json import dumps
from os import environ

class Mock:

    def __getattribute__(self, attr):
        try:
            return object.__getattribute__(self, attr)
        except AttributeError:
            return ""
class Args:

    def __init__(self, args: parser):
        filepath = args.filepath
        filename = args.filename
        path_to_dump_excel = args.excel_output
        template = args.template

args = Args(parser()) if environ.get("CLI") else Mock()

logging.basicConfig(format="%(levelname)s - %(message)s", level=logging.ERROR)    
def list_to_str_serializer(lst: list[str]) -> str:
    try:
        tmp = "".join(lst)
    except (ValueError, AttributeError):
        tmp = ""
        logging.info("Serializing arguments...")
    return tmp
        
def filename_handler(filename: str = list_to_str_serializer(args.filename)) -> load_workbook:
    if not check_xlsx_format(filename):
        dir, tail = os.path.split(filename)
        if not dir:
            dir = None
        _filename = "".join([file for file in os.listdir(dir) if not os.path.isdir(file) and tail == file[:file.rfind(".")]])
        if _filename:
            filename = os.path.join(dir, _filename) if dir else _filename
    try:
        workbook = load_workbook(filename)
    except (FileNotFoundError, InvalidFileException) as exception:
        if not os.path.split(filename)[1] == "sheets":
            filename = os.path.join("sheets", filename)
            return filename_handler(filename)
        if isinstance(exception, FileNotFoundError):
            logging.error("Whoops..Can't find this file")
        elif isinstance(exception, InvalidFileException):
            logging.error("Whoops..Supported formats are: .xlsx,.xlsm,xltx,.xltm")
    except BaseException as exc:
        logging.error(f"Unknown error while loading workbook exc: {exc}")
    else:
        return workbook    

    
def check_xlsx_format(filename: str) -> bool:
    return filename.endswith((".xlsx", ".xlsm", ".xltx", ".xltm"))


def collect_excel_files(filepath: str) -> list[str] | None:
    files_result = []
    for (root, _, files) in os.walk(filepath):
        for file in files:
             if check_xlsx_format(file): 
                 files_result.append(os.path.join(root, file))
    return files_result
                 
def filepath_handler(filepath: str = list_to_str_serializer(args.filepath)) -> list[str] | None | str:
    if os.path.isdir(filepath):
        if os.path.isabs(filepath):
            files_result = collect_excel_files(filepath)
            return files_result
        else:
            files_result = collect_excel_files(os.path.join(os.getcwd(), filepath))
            return files_result
    if os.path.isfile(filepath):
        return filename_handler(filepath)
    else:
        filepath = os.path.join(os.getcwd(), "sheets")
        files_result = collect_excel_files(filepath)
        return files_result + [os.path.join(os.getcwd(), file) for file in os.listdir() if check_xlsx_format(file)]


def check_and_update_file_format(filename, format="xlsx"):
    match format:
        case "xlsx": 
            if not check_xlsx_format(filename):
                return f"{filename}.xlsx"
        case "json":
            if not check_json_format(filename):
                return f"{filename}.json"
            
def check_json_format(filename):
    return filename.endswith("json")

def direct_output_to_file(
    dirname: str, 
    filename: str = "example.xlsx", 
    output: any = "", 
    format: Literal["xlsx", "json"] = "xlsx"
    ):
    if not filename:
        filename=f"example.{format}"
    if (_file_name := check_and_update_file_format(filename, format)) is not None: 
        filename = _file_name
            
        
    try:
        with open(os.path.join(dirname, filename), "x") as file:
                    file.write(output)
    except FileExistsError:
        with open(os.path.join(dirname, filename), "w") as file:
            file.write(output)

def create_template(path: str = args.template) -> None:
    if path is None:
        return
    if not path:
        dirname= "templates"
        os.makedirs(dirname, exist_ok=True)
        direct_output_to_file(dirname)
    else:
        head, tail = os.path.split(list_to_str_serializer(path))
        if not head:
            head, tail = tail, None
        os.makedirs(head, exist_ok=True)
        direct_output_to_file(head, tail)
        
def workbook_serializer(wb: Workbook):
    worksheet = wb.active
    cell_identifiers = [(cell.row, cell.column, cell.value) for cell in worksheet['1'] if cell.value is not None]
    result = []
    table_index = 1
    while True:
        tmp = dict()
        for row_index, column_index, key in cell_identifiers:
            cell = worksheet.cell(row_index + table_index, column_index)
            tmp[key] = cell.value
        if len(set(tmp.values())) <= 1:
            return result
        else:
            table_index += 1
            result.append(tmp)

def run_parser() -> list[dict]:
    if args.filename and (workbook := filename_handler()):
        return workbook_serializer(workbook)
    excel_files = filepath_handler()
    if isinstance(excel_files, Workbook):
        return workbook_serializer(excel_files)
    worksheets: list[Workbook] = [filename_handler(wb) for wb in excel_files]
    user_data = sum([workbook_serializer(worksheet) for worksheet in worksheets], []) # Removing arrays nesting
    return user_data


def write_excel_output(data_to_dump, path: str = list_to_str_serializer(args.path_to_dump_excel)) -> None:
    if type(data_to_dump) is list:
        data_to_dump = dumps(data_to_dump, indent=5)
    if not path:
        path = "DATA"
    head, tail = os.path.split(list_to_str_serializer(path))
    if not head: head, tail = tail, None
    os.makedirs(head, exist_ok=True)
    direct_output_to_file(head, tail, data_to_dump, format="json")