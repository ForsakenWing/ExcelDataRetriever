from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from .cli import Args
import os
import logging

logging.basicConfig(format="%(levelname)s - %(message)s", level=logging.INFO)
def filename_handler(filename: str = Args.filename) -> load_workbook:
    try:
        workbook = load_workbook(filename)
    except FileNotFoundError:
        logging.error("Whoops..Can't find this file")
    except InvalidFileException:
        logging.error("Whoops..Supported formats are: .xlsx,.xlsm,xltx,.xltm")
    else:
        return workbook    
    
def check_right_format(filename: str) -> bool:
    return filename.endswith((".xlsx", ".xlsm", ".xltx", ".xltm"))

def collect_excel_files(filepath: str) -> list[str] | None:
    files_result = []
    for (root, _, files) in os.walk(filepath):
        for file in files:
             if check_right_format(file): 
                 files_result.append(os.path.join(root, file))
    return files_result
                 
def filepath_handler(filepath: str = Args.filepath) -> list[str] | None | str:
    if os.path.isdir(filepath):
        if os.path.isabs(filepath):
            files_result = collect_excel_files(filepath)
            return files_result
        else:
            files_result = collect_excel_files(os.path.join(os.getcwd(), filepath))
            return files_result
    if os.path.isfile(filepath):
        return filename_handler(filepath)
    else:
        filepath = os.path.join(os.getcwd(), "sheets")
        files_result = collect_excel_files(filepath)
        return files_result + [os.path.join(os.getcwd(), file) for file in os.listdir() if check_right_format(file)]

def create_template(path: str = Args.template) -> None:
    if path is not None and not path:
        try:
            dirname= "templates"
            os.mkdir(dirname)
        except OSError:
            logging.info("templates directory already exists")
        finally:
            try:
                filename = "example.xlsx"
                with open(os.path.join(dirname, filename), "x") as file:
                    file.write("MOCK")
            except FileExistsError:
                with open(os.path.join(dirname, filename), "w") as file:
                    file.write("MOCK")

